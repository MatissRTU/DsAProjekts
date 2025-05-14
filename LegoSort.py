import requests
from bs4 import BeautifulSoup
import openpyxl

class Node: 
	def __init__(self, key, value): 
		self.key = key 
		self.value = value 
		self.next = None

#hash implement
class HashTable: 
	def __init__(self, capacity): 
		self.capacity = capacity 
		self.size = 0
		self.table = [None] * capacity 

	def _hash(self, key): 
		return hash(key) % self.capacity 

	def insert(self, key, value): 
		index = self._hash(key) 

		if self.table[index] is None: 
			self.table[index] = Node(key, value) 
			self.size += 1
		else: 
			current = self.table[index] 
			while current: 
				if current.key == key: 
					if isinstance(current.value, list) and isinstance(current.value[0], list):
						current.value.append(value)
					else:
						current.value = [current.value, value]
					return
				current = current.next
			new_node = Node(key, value) 
			new_node.next = self.table[index] 
			self.table[index] = new_node 
			self.size += 1

	def search(self, key): 
		index = self._hash(key) 

		current = self.table[index] 
		while current: 
			if current.key == key: 
				return current.value 
			current = current.next

		raise KeyError(key) 

	def remove(self, key): 
		index = self._hash(key) 

		previous = None
		current = self.table[index] 

		while current: 
			if current.key == key: 
				if previous: 
					previous.next = current.next
				else: 
					self.table[index] = current.next
				self.size -= 1
				return
			previous = current 
			current = current.next

		raise KeyError(key) 

	def __len__(self): 
		return self.size 

	def __contains__(self, key): 
		try: 
			self.search(key) 
			return True
		except KeyError: 
			return False
#hash implement end	

# ~~~~~~~~~~~~~~~~~~	
#           _
#       .__(.)< (MEOW)
#        \___)   
# ~~~~~~~~~~~~~~~~~~

def search1(url,id):#prieks 1alv
	page = requests.get(url, headers=id)
	if page.status_code == 200:
		soup = BeautifulSoup(page.content, "html.parser")
		pagination = soup.select_one(".catalog-taxons-pagination .paginator__last")  # elements pēdējam lapas ciparam
		last_page = int(pagination.text.strip())  # atdala ciparu no elementa
	
		for page_number in range(1,1+1):#KAD TESTE last_page samainit ar 1
			search_url = f"{url}&page={page_number}"
			page = requests.get(search_url, headers=id)

			soup = BeautifulSoup(page.content, "html.parser")
			product_sections = soup.select("div.catalog-taxons-product")
			print(f"searching({page_number}/{last_page})...")

			for block in product_sections:
				class_list = block.get("class", [])#nolasa klases ipasibas
				if "catalog-taxons-product--no-product" in class_list:#objekts nosaka ka produkts izpardots
					continue
            
				itemdata = block.find("div", class_="gtm-categories")
				itemimg = block.find("img", class_="catalog-taxons-product__image")

				if itemimg: 
					img = itemimg.get("data-src") or itemimg.get("src")
				if itemdata:
					name = itemdata.get("data-name")
					price = itemdata.get("data-price")
					product_data.insert(round(float(price),2),[name, img])

def search2(url,id):# prieks m79
    page = 1

    while True:
        print(f"Searching page {page}...")
        page_url = url.replace("/1/", f"/{page}/")
        response = requests.get(page_url,headers=id)

        if response.status_code != 200:
            print(f"lapai nevar piekļūt, kods:{response.status_code}")
            break

        soup = BeautifulSoup(response.text, "html.parser")
        items = soup.find_all("div", class_="item")
        if len(items) == 0:
            print("Beidzas produkti")
            break

        for item in items:
            img_tag = item.find("img")
            image_link = img_tag['src'] if img_tag else None

            name = img_tag.get("alt") if img_tag and img_tag.has_attr("alt") else "No Name"

            price_block = item.find(class_="price")
            price = None

            if price_block:
                price_tag = price_block.find("b", itemprop="price")
                span_tag = price_block.find("span")
                if price_tag and span_tag:
                    try:
                        price = float(price_tag.text.strip().replace(",", "."))
                    except ValueError:
                        continue

            if image_link and price is not None:
                product_data.insert(round(price, 2), [name, image_link])

        page += 1


def sort_to_excel(price_range):
	try:
		Excel = openpyxl.Workbook()
		doc = Excel.active  # atver Excel
		doc.title = "LEGO komplektu akcijas buklets"
		doc.append(["Nosaukums", "Cena", "Bilde"])
		### SEIT VEIKT FILTRESANU

		doc.column_dimensions['A'].width = 80  # Name
		doc.column_dimensions['B'].width = 10  # Price
		doc.column_dimensions['C'].width = 18  # Image URL (just for reference)

		for block in product_data.table:
			current = block
			while current:
				key = current.key
				if key <= price_range:
					value = current.value
					if isinstance(value, list) and all(isinstance(i, list) and len(i) == 2 for i in value):
						for item in value:
							name, url = item
							doc.append([name, f"{key}€", f'=_xlfn.IMAGE("{url}")'])
							doc.row_dimensions[doc.max_row].height = 100  # Set cell height for image

					# Handle a single [name, url] entry
					else:
						name, url = value
						doc.append([name, f"{key}€", f'=_xlfn.IMAGE("{url}")'])
						doc.row_dimensions[doc.max_row].height = 100
				current = current.next

		Excel.save("Lego_akcijas.xlsx")
		print("Excel save complete")
	except Exception:
		print(f"Fails nav aizvērts, aizvērt to un restartēt programmu")

#Main programmas sakums
userid = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"#nepieciesams ksenukajam
}

#izmantojamie url meklesana
url1 = "https://www.1a.lv/c/berniem-mazuliem/lego-rotallietas-un-lelles/lego/37h?lf=1"#                           |
url2 = "https://m79.lv/rotallietasspeles/lego/1/f:5637150982-5637212832" 


product_data = HashTable(5000)

print("Sākta meklēšana 1A.lv...")
search1(url1,userid)
print("Sākta meklēšana m79.lv...")
search2(url2,userid)

while True:
	tester = input("Ievadīt maksimālo cenu vai exit ,lai izslēgtu programmu: ")
	if tester == "exit":
		break
	else:
		sort_to_excel(float(tester))
