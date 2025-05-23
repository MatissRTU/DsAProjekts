import requests
from bs4 import BeautifulSoup
import openpyxl

class Node: 
	def __init__(self, key, value): 
		self.key = key 
		self.value = value 
		self.next = None

#hash implement
class HashTable: 
	def __init__(self, capacity): 
		self.capacity = capacity 
		self.size = 0
		self.table = [None] * capacity 

	def _hash(self, key): 
		return hash(key) % self.capacity 

	def insert(self, key, value): 
		index = self._hash(key) 

		if self.table[index] is None: 
			self.table[index] = Node(key, value) 
			self.size += 1
		else: 
			current = self.table[index] 
			while current: 
				if current.key == key: 
					if isinstance(current.value, list) and isinstance(current.value[0], list):
						current.value.append(value)
					else:
						current.value = [current.value, value]
					return
				current = current.next
			new_node = Node(key, value) 
			new_node.next = self.table[index] 
			self.table[index] = new_node 
			self.size += 1

	def search(self, key): 
		index = self._hash(key) 

		current = self.table[index] 
		while current: 
			if current.key == key: 
				return current.value 
			current = current.next

		raise KeyError(key) 

	def remove(self, key): 
		index = self._hash(key) 

		previous = None
		current = self.table[index] 

		while current: 
			if current.key == key: 
				if previous: 
					previous.next = current.next
				else: 
					self.table[index] = current.next
				self.size -= 1
				return
			previous = current 
			current = current.next

		raise KeyError(key) 

	def __len__(self): 
		return self.size 

	def __contains__(self, key): 
		try: 
			self.search(key) 
			return True
		except KeyError: 
			return False
#hash implement end	


def search1(url,id):#prieks ksenukai/1alv
	page = requests.get(url, headers=id)
	if page.status_code == 200:
		soup = BeautifulSoup(page.content, "html.parser")
		pagination = soup.select_one(".catalog-taxons-pagination .paginator__last")  # elements pēdējam lapas ciparam
		last_page = int(pagination.text.strip())  # atdala ciparu no elementa
	
		for page_number in range(1,1+1):#KAD TESTE last_page samainit ar 1
			search_url = f"{url}&page={page_number}"
			page = requests.get(search_url, headers=id)

			soup = BeautifulSoup(page.content, "html.parser")
			product_sections = soup.select("div.catalog-taxons-product")
			print(f"searching({page_number}/{last_page})...")

			for block in product_sections:
				class_list = block.get("class", [])#nolasa klases ipasibas
				if "catalog-taxons-product--no-product" in class_list:#objekts nosaka ka produkts izpardots
					continue
            
				itemdata = block.find("div", class_="gtm-categories")
				itemimg = block.find("img", class_="catalog-taxons-product__image")

				if itemimg: 
					img = itemimg.get("data-src") or itemimg.get("src")
				if itemdata:
					name = itemdata.get("data-name")
					price = itemdata.get("data-price")
					product_data.insert(round(float(price),2),[name, img])

def search2(url,id):# TODO remake to 220lv
    page = 1

    while True:
        print(f"Searching page {page}...")
        page_url = url.replace("/1/", f"/{page}/")
        response = requests.get(page_url,headers=id)

        if response.status_code != 200:
            print(f"lapai nevar piekļūt, kods:{response.status_code}")
            break
        print("test1")   ######     
        soup = BeautifulSoup(response.text, "html.parser")
        items = soup.find_all("div", class_="item")
        if len(items) == 0:
            print("No items found. Stopping search.")
            break

        print("test2") ######  
        for item in items:
            print(f"Searching item in page {page}...")  # Debug: Should print this for each item
            # Extract image
            img_tag = item.find("img")
            image_link = img_tag['src'] if img_tag else None

            # Try to get product name (fallback to alt text or empty string)
            name = img_tag.get("alt") if img_tag and img_tag.has_attr("alt") else "No Name"

            # Get price and value
            price_block = item.find(class_="price")
            price = None

            if price_block:
                print(f"Searching price in item {page}...")  # Debug: Should print this when price block found
                price_tag = price_block.find("b", itemprop="price")
                span_tag = price_block.find("span")
                if price_tag and span_tag:
                    try:
                        price = float(price_tag.text.strip().replace(",", "."))
                    except ValueError:
                        continue  # skip if price isn't a valid number

            if image_link and price is not None:
                print(f"Name: {name}, Price: {price}, Image: {image_link}")
                product_data.insert(round(price, 2), [name, image_link])

        page += 1


def sort_to_excel(price_range):
	try:
		Excel = openpyxl.Workbook()
		doc = Excel.active  # atver Excel
		doc.title = "LEGO komplektu akcijas buklets"
		doc.append(["Nosaukums", "Cena", "Bilde"])
		### SEIT VEIKT FILTRESANU

		doc.column_dimensions['A'].width = 80  # Name
		doc.column_dimensions['B'].width = 10  # Price
		doc.column_dimensions['C'].width = 18  # Image URL (just for reference)

		for block in product_data.table:
			current = block
			while current:
				key = current.key
				if key <= price_range:
					value = current.value
					if isinstance(value, list) and all(isinstance(i, list) and len(i) == 2 for i in value):
						for item in value:
							name, url = item
							doc.append([name, f"{key}€", f'=_xlfn.IMAGE("{url}")'])
							doc.row_dimensions[doc.max_row].height = 100  # Set cell height for image

					# Handle a single [name, url] entry
					else:
						name, url = value
						doc.append([name, f"{key}€", f'=_xlfn.IMAGE("{url}")'])
						doc.row_dimensions[doc.max_row].height = 100

				current = current.next

		Excel.save("Lego_akcijas.xlsx")
		print("Excel save complete")
	except Exception:
		print(f"Fails nav aizvērts, aizvērt to un restartēt programmu")

userid = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"#nepieciesams ksenukajam
}
userid2 = {
	"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.6312.105 Safari/537.36"
}

#izmantojamie url meklesana
url1 = "https://www.1a.lv/c/berniem-mazuliem/lego-rotallietas-un-lelles/lego/37h?lf=1"#                                                              
url2 = "https://m79.lv/rotallietasspeles/lego/1/f:5637150982-5637212832" 


def test_values():
	product_data.insert(10, ["Test Product 1", "https://cdn.discordapp.com/attachments/828739920633790525/1370267347922976768/b6cf7efbc832a7f3c9bf8c3305af0afb.png?ex=68237daa&is=68222c2a&hm=075fe3902c88637160e0f267dae12ac5f1488427a69789eea586bbe0696678c7&"])
	product_data.insert(-20, ["Test Product 2", "https://media.discordapp.net/attachments/793959529812721704/1367226777197482095/image.png?ex=6822f9e9&is=6821a869&hm=c7539c40b690006efb19b7ae52c4eb70bbb1fc4ea3a237e1b0089cbb554973c4&=&format=webp&quality=lossless&width=142&height=129"])
	product_data.insert(666, ["Test Product 3", "https://cdn.discordapp.com/attachments/995490453103317132/1357840486743801998/image.png?ex=68231b3f&is=6821c9bf&hm=63b6715bbd17a8c2c45444b71d2551c57d6c0fe406ff845cebb33df21c2eeebc&"])
	product_data.insert(5, ["Test Product 4", "https://cdn.discordapp.com/attachments/828739920633790525/1368476463300350042/5597c9857a405a83bddff0a723d9e6eb.png?ex=68239145&is=68223fc5&hm=4cad155a87311134a70bd762e282b3a062428b683ea03b710b0a0ff738b35231&"])
	product_data.insert(5, ["Test Product 5", "https://cdn.discordapp.com/attachments/995490453103317132/1334511314474897519/20250130_151003.jpg?ex=68234503&is=6821f383&hm=041d28ea39b3d1f50c9f15c99927c662c58e6954cdd1bd7ef0cc192e612b0c04&"])

product_data = HashTable(6000)

#search1(url1,userid)#TODO FINISH AND UNCOMMENT
search2(url2,userid)

#test_values()
#sort_to_excel(500)
#page = requests.get(url4,headers=userid)
#print(page.status_code)
